import datetime
from pathlib import Path

import xlsxwriter
from openpyxl import load_workbook


def open_excel(name_file):  # функция получения сырого массива из файла
    # НАДО ДОБАВИТЬ ИСКЛЮЧЕНИЯ ЕСЛИ ФАЙЛА НЕТ
    wb = load_workbook(filename=name_file, read_only=True)
    ws = wb['Лист1']
    # массив для данных из excel
    mass_excel = []
    i = 0
    for row in ws.rows:
        # добавить строку в массив
        mass_excel.append([])
        for cell in row:
            # добавить содержимое ячейки в массив
            mass_excel[i].append(cell.value)
            # cell.row = 100   можно изменить строку
        i += 1
    # закрытие документа
    wb.close()

    # вернуть массив
    return mass_excel


def f_save_xlsx(name_file: str, dir: str, header: list, data: list):
    data.insert(0, header)

    now = datetime.datetime.now()
    name_f = f'{name_file} ' + str(now).replace(':', '_') + '.xlsx'

    path_for_file = dir

    if Path(path_for_file).exists() is False:
        path_dir = Path(path_for_file)
        path_dir.mkdir()

    name_file = Path(path_for_file, name_f)

    # Create a workbook and add a worksheet.
    workbook = xlsxwriter.Workbook(name_file)
    worksheet = workbook.add_worksheet()

    # Start from the first cell. Rows and columns are zero indexed.
    row = 0

    # Iterate over the data and write it out row by row.
    for line_r in data:
        col = 0
        for line_c in line_r:
            worksheet.write(row, col, line_c)
            col += 1
        row += 1

    workbook.close()

    print(f'save file comlite: {name_f}')
